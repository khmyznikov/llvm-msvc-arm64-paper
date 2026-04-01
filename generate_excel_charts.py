"""
Generate Excel workbook with performance comparison charts:
  LLVM (clang-cl) vs MSVC on two ARM64 machines.
"""

import json, math, statistics
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MACHINES = {
    "gleb-devkit-arm64": "Snapdragon 8cx Gen 3 (Volterra)",
    "gleb-surface-15-arm64": "Surface Laptop 15 (X Plus)",
}
RESULTS = Path("results")

# ── colours ──
LLVM_FILL = "4472C4"   # blue
MSVC_FILL = "ED7D31"   # orange
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def load_json(path):
    with open(path) as f:
        return json.load(f)


# ────────────────────────── data extraction ──────────────────────────

def get_lame_data():
    """Return {machine: {toolchain: mean_sec}}."""
    data = {}
    for mid in MACHINES:
        data[mid] = {}
        for tc in ("llvm", "msvc"):
            d = load_json(RESULTS / mid / "lame" / f"lame_{tc}_arm64.json")
            data[mid][tc] = d["mean_sec"]
    return data


def get_x264_data():
    """Return {machine: {toolchain: mean_fps}}."""
    data = {}
    for mid in MACHINES:
        d = load_json(RESULTS / mid / "x264" / "x264_results_arm64.json")
        data[mid] = {
            "msvc": d["results"]["msvc"]["mean_fps"],
            "llvm": d["results"]["llvm"]["mean_fps"],
        }
    return data


def get_numpy_data():
    """Return {machine: {toolchain: {op_name: mean_sec, ...}}}."""
    data = {}
    for mid in MACHINES:
        data[mid] = {}
        for tc in ("llvm", "msvc"):
            d = load_json(RESULTS / mid / "numpy" / f"numpy_{tc}_arm64.json")
            # Support both old single-op format and new multi-op format
            if "operations" in d:
                ops = {}
                for op_name, op_data in d["operations"].items():
                    ops[op_name] = op_data["mean_sec"]
                data[mid][tc] = ops
            else:
                # Legacy format: single count_nonzero result
                data[mid][tc] = {"count_nonzero": statistics.mean(d["times_sec"])}
    return data


def numpy_geo_mean(numpy, mid, tc):
    """Geometric mean of all numpy operations."""
    vals = list(numpy[mid][tc].values())
    return math.exp(sum(math.log(v) for v in vals) / len(vals))


def get_cpython_data():
    """Return {machine: {toolchain: {bench_name: mean_sec, '__geo__': val}}}."""
    data = {}
    for mid in MACHINES:
        data[mid] = {}
        for tc in ("llvm", "msvc"):
            path = RESULTS / mid / "cpython" / f"pyperformance_{tc}_arm64.json"
            d = load_json(path)
            benchmarks = {}
            for b in d["benchmarks"]:
                name = b["metadata"]["name"]
                values = []
                for run in b["runs"]:
                    if "values" in run:
                        values.extend(run["values"])
                if values:
                    benchmarks[name] = statistics.mean(values)
            geo = math.exp(
                sum(math.log(v) for v in benchmarks.values()) / len(benchmarks)
            )
            benchmarks["__geo__"] = geo
            data[mid][tc] = benchmarks
    return data


# ────────────────────────── helpers ──────────────────────────

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, size=11)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 3


def pct_diff(llvm, msvc, higher_is_better=False):
    """Return LLVM advantage as a percentage string."""
    if higher_is_better:
        diff = (llvm - msvc) / msvc * 100
    else:
        diff = (msvc - llvm) / msvc * 100
    return f"{diff:+.1f}%"


def add_bar_chart(ws, title, min_row, max_row, min_col, max_col, cat_col,
                  y_axis_title, position="E2", width=22, height=14,
                  higher_is_better=False):
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = None
    chart.width = width
    chart.height = height
    chart.legend.position = "b"

    cats = Reference(ws, min_col=cat_col, min_row=min_row + 1, max_row=max_row)
    for col_idx in range(min_col, max_col + 1):
        values = Reference(ws, min_col=col_idx, min_row=min_row, max_row=max_row)
        chart.add_data(values, titles_from_data=True)
    chart.set_categories(cats)

    # colour series
    if len(chart.series) >= 1:
        chart.series[0].graphicalProperties.solidFill = LLVM_FILL
    if len(chart.series) >= 2:
        chart.series[1].graphicalProperties.solidFill = MSVC_FILL

    # data labels
    for s in chart.series:
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.numFmt = '0.00' if not higher_is_better else '0.00'

    ws.add_chart(chart, position)
    return chart


# ────────────────────────── worksheets ──────────────────────────

def sheet_overview(wb, lame, x264, numpy, cpython):
    """Summary sheet: one row per benchmark, columns per machine × toolchain."""
    ws = wb.active
    ws.title = "Overview"

    headers = ["Benchmark", "Metric"]
    for mid, label in MACHINES.items():
        headers += [f"{label}\nLLVM", f"{label}\nMSVC", f"{label}\nΔ LLVM"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    # LAME – lower is better (seconds)
    row = ["LAME MP3 encode", "Time (s)"]
    for mid in MACHINES:
        row += [round(lame[mid]["llvm"], 4), round(lame[mid]["msvc"], 4),
                pct_diff(lame[mid]["llvm"], lame[mid]["msvc"])]
    ws.append(row)

    # x264 – higher is better (fps)
    row = ["x264 H.264 encode", "FPS"]
    for mid in MACHINES:
        row += [round(x264[mid]["llvm"], 2), round(x264[mid]["msvc"], 2),
                pct_diff(x264[mid]["llvm"], x264[mid]["msvc"], higher_is_better=True)]
    ws.append(row)

    # NumPy – geo-mean of compiler-sensitive ops, lower is better
    row = ["NumPy (sqrt, sort)", "Geo-mean (ms)"]
    for mid in MACHINES:
        llvm_geo = numpy_geo_mean(numpy, mid, "llvm")
        msvc_geo = numpy_geo_mean(numpy, mid, "msvc")
        row += [round(llvm_geo * 1000, 4), round(msvc_geo * 1000, 4),
                pct_diff(llvm_geo, msvc_geo)]
    ws.append(row)

    # CPython geometric mean – lower is better
    row = ["CPython pyperformance", "Geo-mean (s)"]
    for mid in MACHINES:
        row += [round(cpython[mid]["llvm"]["__geo__"], 6),
                round(cpython[mid]["msvc"]["__geo__"], 6),
                pct_diff(cpython[mid]["llvm"]["__geo__"], cpython[mid]["msvc"]["__geo__"])]
    ws.append(row)

    auto_width(ws)

    # ── Overview chart 1: Absolute performance (grouped by benchmark, LLVM vs MSVC × machine) ──
    chart_ws = wb.create_sheet("Overview Chart")

    # Build a normalized performance table (higher = better for all benchmarks)
    # For time-based: invert to throughput. For fps-based: use directly.
    machine_labels = list(MACHINES.values())
    chart_ws.append(["Configuration", "LAME MP3\n(encodes/s)", "x264\n(fps)",
                     "NumPy sort\n(1/time)", "CPython\n(1/geo-mean)"])
    style_header(chart_ws, 1, 5)

    for mid, label in MACHINES.items():
        # Use sort as the most dramatic compiler-sensitive NumPy op
        np_sort_llvm = numpy[mid]["llvm"].get("sort", 0.001)
        np_sort_msvc = numpy[mid]["msvc"].get("sort", 0.001)
        chart_ws.append([
            f"{label} — LLVM",
            round(1.0 / lame[mid]["llvm"], 3),
            round(x264[mid]["llvm"], 2),
            round(1.0 / np_sort_llvm, 2),
            round(1.0 / cpython[mid]["llvm"]["__geo__"], 2),
        ])
        chart_ws.append([
            f"{label} — MSVC",
            round(1.0 / lame[mid]["msvc"], 3),
            round(x264[mid]["msvc"], 2),
            round(1.0 / np_sort_msvc, 2),
            round(1.0 / cpython[mid]["msvc"]["__geo__"], 2),
        ])

    # One chart per benchmark column so scales don't clash
    bench_cols = [
        (2, "LAME MP3 — Throughput (encodes/s)", "encodes/s"),
        (3, "x264 — Encode FPS", "fps"),
        (4, "NumPy sort — Throughput (sorts/s)", "sorts/s"),
        (5, "CPython pyperformance — 1/geo-mean (higher=better)", "1/geo-mean"),
    ]
    positions = ["A8", "N8", "A26", "N26"]
    bar_colors = [LLVM_FILL, MSVC_FILL, LLVM_FILL, MSVC_FILL]  # alternating per config row

    for (col_idx, title, yaxis), pos in zip(bench_cols, positions):
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = title
        chart.y_axis.title = yaxis
        chart.width = 22
        chart.height = 14
        chart.legend.position = "b"

        cats = Reference(chart_ws, min_col=1, min_row=2, max_row=5)
        vals = Reference(chart_ws, min_col=col_idx, min_row=1, max_row=5)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)

        # Colour bars: row1=machine1-LLVM, row2=machine1-MSVC, row3=machine2-LLVM, row4=machine2-MSVC
        from openpyxl.chart.series import DataPoint
        from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
        series = chart.series[0]
        colors = [LLVM_FILL, MSVC_FILL, LLVM_FILL, MSVC_FILL]
        for i, c in enumerate(colors):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = c
            series.data_points.append(pt)

        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.numFmt = '0.00'

        chart_ws.add_chart(chart, pos)

    # ── Overview chart 2: LLVM advantage % ──
    adv_ws = wb.create_sheet("LLVM Advantage %")
    adv_ws.append(["Benchmark"] + list(MACHINES.values()))
    style_header(adv_ws, 1, 1 + len(MACHINES))

    benchmarks_for_chart = [
        ("LAME MP3", lambda mid: (lame[mid]["msvc"] - lame[mid]["llvm"]) / lame[mid]["msvc"] * 100),
        ("x264", lambda mid: (x264[mid]["llvm"] - x264[mid]["msvc"]) / x264[mid]["msvc"] * 100),
        ("NumPy", lambda mid: (numpy_geo_mean(numpy, mid, "msvc") - numpy_geo_mean(numpy, mid, "llvm")) / numpy_geo_mean(numpy, mid, "msvc") * 100),
        ("CPython", lambda mid: (cpython[mid]["msvc"]["__geo__"] - cpython[mid]["llvm"]["__geo__"]) / cpython[mid]["msvc"]["__geo__"] * 100),
    ]
    for name, fn in benchmarks_for_chart:
        row = [name]
        for mid in MACHINES:
            row.append(round(fn(mid), 1))
        adv_ws.append(row)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "LLVM Advantage Over MSVC (%)"
    chart.y_axis.title = "LLVM faster by (%)"
    chart.width = 24
    chart.height = 14
    chart.legend.position = "b"

    cats = Reference(adv_ws, min_col=1, min_row=2, max_row=5)
    for ci in range(2, 2 + len(MACHINES)):
        vals = Reference(adv_ws, min_col=ci, min_row=1, max_row=5)
        chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.solidFill = "4472C4"
    chart.series[1].graphicalProperties.solidFill = "70AD47"

    for s in chart.series:
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.numFmt = '0.0"%"'

    adv_ws.add_chart(chart, "A8")
    auto_width(adv_ws)


def sheet_lame(wb, lame):
    ws = wb.create_sheet("LAME MP3")
    ws.append(["Machine", "LLVM (s)", "MSVC (s)", "LLVM Advantage"])
    style_header(ws, 1, 4)
    for mid, label in MACHINES.items():
        ws.append([label, round(lame[mid]["llvm"], 4), round(lame[mid]["msvc"], 4),
                    pct_diff(lame[mid]["llvm"], lame[mid]["msvc"])])

    add_bar_chart(ws, "LAME MP3 Encode Time", 1, 3, 2, 3, 1,
                  "Time (seconds)", position="A6")
    auto_width(ws)


def sheet_x264(wb, x264):
    ws = wb.create_sheet("x264")
    ws.append(["Machine", "LLVM (fps)", "MSVC (fps)", "LLVM Advantage"])
    style_header(ws, 1, 4)
    for mid, label in MACHINES.items():
        ws.append([label, round(x264[mid]["llvm"], 2), round(x264[mid]["msvc"], 2),
                    pct_diff(x264[mid]["llvm"], x264[mid]["msvc"], higher_is_better=True)])

    add_bar_chart(ws, "x264 Encode FPS (higher is better)", 1, 3, 2, 3, 1,
                  "Frames per second", position="A6", higher_is_better=True)
    auto_width(ws)


def sheet_numpy(wb, numpy):
    """Single NumPy sheet: per-operation rows for each machine, with chart."""
    ws = wb.create_sheet("NumPy")

    headers = ["Machine", "Operation", "LLVM (ms)", "MSVC (ms)", "LLVM Advantage"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    row_count = 0
    for mid, label in MACHINES.items():
        # Use each machine's own operation list
        op_names = list(numpy[mid]["llvm"].keys())
        for op in op_names:
            llvm_v = numpy[mid]["llvm"].get(op, 0)
            msvc_v = numpy[mid]["msvc"].get(op, 0)
            ws.append([label, op,
                        round(llvm_v * 1000, 4),
                        round(msvc_v * 1000, 4),
                        pct_diff(llvm_v, msvc_v) if msvc_v else ""])
            row_count += 1
        # geo-mean row per machine
        llvm_geo = numpy_geo_mean(numpy, mid, "llvm")
        msvc_geo = numpy_geo_mean(numpy, mid, "msvc")
        ws.append([label, "GEOMETRIC MEAN",
                    round(llvm_geo * 1000, 4),
                    round(msvc_geo * 1000, 4),
                    pct_diff(llvm_geo, msvc_geo)])
        row_count += 1

    nrows = 1 + row_count  # header + data rows
    add_bar_chart(ws, "NumPy Operations — LLVM vs MSVC",
                  1, nrows, 3, 4, 2,
                  "Time (ms)", position="G2", width=28, height=16)
    auto_width(ws)


def sheet_cpython(wb, cpython):
    """One sheet per machine with per-benchmark bars + a combined geo-mean sheet."""
    bench_names = [n for n in cpython[list(MACHINES.keys())[0]]["llvm"] if n != "__geo__"]

    for mid, label in MACHINES.items():
        ws = wb.create_sheet(f"CPython - {label[:20]}")
        ws.append(["Benchmark", "LLVM (s)", "MSVC (s)", "LLVM Advantage"])
        style_header(ws, 1, 4)

        for name in bench_names:
            llvm_v = cpython[mid]["llvm"].get(name, 0)
            msvc_v = cpython[mid]["msvc"].get(name, 0)
            ws.append([name, round(llvm_v, 6), round(msvc_v, 6),
                        pct_diff(llvm_v, msvc_v) if msvc_v else ""])

        # geo mean row
        ws.append(["GEOMETRIC MEAN",
                    round(cpython[mid]["llvm"]["__geo__"], 6),
                    round(cpython[mid]["msvc"]["__geo__"], 6),
                    pct_diff(cpython[mid]["llvm"]["__geo__"], cpython[mid]["msvc"]["__geo__"])])

        nrows = len(bench_names) + 2  # header + data + geo
        add_bar_chart(ws, f"CPython pyperformance — {label}",
                      1, nrows, 2, 3, 1,
                      "Time (seconds)", position="F2", width=28, height=18)
        auto_width(ws)

    # ── Combined geo-mean comparison ──
    ws = wb.create_sheet("CPython Geo-Mean")
    ws.append(["Machine", "LLVM (s)", "MSVC (s)", "LLVM Advantage"])
    style_header(ws, 1, 4)
    for mid, label in MACHINES.items():
        ws.append([label,
                    round(cpython[mid]["llvm"]["__geo__"], 6),
                    round(cpython[mid]["msvc"]["__geo__"], 6),
                    pct_diff(cpython[mid]["llvm"]["__geo__"], cpython[mid]["msvc"]["__geo__"])])

    add_bar_chart(ws, "CPython pyperformance Geometric Mean", 1, 3, 2, 3, 1,
                  "Geometric mean time (s)", position="A6")
    auto_width(ws)


# ────────────────────────── LLVM speedup ratio sheet ──────────────────────────

def sheet_speedup(wb, lame, x264, numpy, cpython):
    """Speedup ratio = MSVC_time / LLVM_time  (>1 means LLVM is faster)."""
    ws = wb.create_sheet("Speedup Ratio")
    ws.append(["Benchmark"] + [MACHINES[m] for m in MACHINES])
    style_header(ws, 1, 1 + len(MACHINES))

    rows = [
        ("LAME MP3", {m: lame[m]["msvc"] / lame[m]["llvm"] for m in MACHINES}),
        ("x264", {m: x264[m]["llvm"] / x264[m]["msvc"] for m in MACHINES}),
        ("NumPy (geo)", {m: numpy_geo_mean(numpy, m, "msvc") / numpy_geo_mean(numpy, m, "llvm") for m in MACHINES}),
        ("CPython (geo)", {m: cpython[m]["msvc"]["__geo__"] / cpython[m]["llvm"]["__geo__"] for m in MACHINES}),
    ]
    for name, vals in rows:
        ws.append([name] + [round(vals[m], 2) for m in MACHINES])

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "LLVM / MSVC Speedup Ratio (>1 = LLVM faster)"
    chart.y_axis.title = "Speedup ratio"
    chart.width = 24
    chart.height = 14
    chart.legend.position = "b"

    cats = Reference(ws, min_col=1, min_row=2, max_row=5)
    for ci in range(2, 2 + len(MACHINES)):
        vals = Reference(ws, min_col=ci, min_row=1, max_row=5)
        chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.solidFill = "4472C4"
    chart.series[1].graphicalProperties.solidFill = "70AD47"

    for s in chart.series:
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.numFmt = '0.00"x"'

    ws.add_chart(chart, "A8")
    auto_width(ws)


# ────────────────────────── main ──────────────────────────

def main():
    lame = get_lame_data()
    x264 = get_x264_data()
    numpy_d = get_numpy_data()
    cpython = get_cpython_data()

    wb = Workbook()

    sheet_overview(wb, lame, x264, numpy_d, cpython)
    sheet_lame(wb, lame)
    sheet_x264(wb, x264)
    sheet_numpy(wb, numpy_d)
    sheet_cpython(wb, cpython)
    sheet_speedup(wb, lame, x264, numpy_d, cpython)

    out = Path("results") / "performance_comparison.xlsx"
    wb.save(out)
    print(f"Saved to {out}")


if __name__ == "__main__":
    main()
